"""
Excel export helper using openpyxl.
"""
from __future__ import annotations

from io import BytesIO
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def records_to_excel(session: dict, records: List[dict]) -> bytes:
    """
    Build an XLSX file in-memory and return raw bytes.

    `session` dict is expected to contain: session_name, class, start_time, end_time, teacher.
    `records` dict list: student_id, student_name, check_in_time, confidence, method.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Header info
    ws["A1"] = "Session Name"
    ws["B1"] = session.get("session_name", "")
    ws["A2"] = "Class"
    ws["B2"] = session.get("class") or session.get("class_name") or ""
    ws["A3"] = "Teacher"
    ws["B3"] = session.get("teacher", "")
    ws["A4"] = "Start Time"
    ws["B4"] = str(session.get("start_time", ""))
    ws["A5"] = "End Time"
    ws["B5"] = str(session.get("end_time", ""))
    ws["A6"] = "Total Attended"
    ws["B6"] = len(records)

    for i in range(1, 7):
        ws[f"A{i}"].font = Font(bold=True)

    # Table headers on row 8
    headers = ["#", "Student ID", "Full Name", "Check-in Time", "Confidence", "Method"]
    header_row = 8
    for idx, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=idx, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="6366F1")  # indigo
        c.alignment = Alignment(horizontal="center")

    for i, rec in enumerate(records, start=1):
        ws.cell(row=header_row + i, column=1, value=i)
        ws.cell(row=header_row + i, column=2, value=rec.get("student_id", ""))
        ws.cell(row=header_row + i, column=3, value=rec.get("student_name", ""))
        ws.cell(row=header_row + i, column=4, value=str(rec.get("check_in_time", "")))
        ws.cell(row=header_row + i, column=5, value=round(float(rec.get("confidence", 0)), 4))
        ws.cell(row=header_row + i, column=6, value=rec.get("method", ""))

    # Column widths
    widths = [6, 14, 30, 22, 12, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
